"""
End-to-end Canopy SFO prospecting pipeline.

Secrets are read from environment variables only. Do not hardcode API keys.

Required environment variables:
  APOLLO_API_KEY
  HUNTER_API_KEY
  SERPER_API_KEY
  BRAVE_API_KEY
  GEMINI_API_KEY

Optional:
  JWT_SECRET
  GEMINI_MODEL, default gemini-1.5-flash-latest

Example:
  python canopy_sfo_pipeline.py --config canopy_sfo_task_config.example.json --out-dir output
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional
from urllib.parse import quote_plus

import requests


TIMEOUT = 30

DEFAULT_CONFIG: Dict[str, Any] = {
    "task": {
        "target_count": 30,
        "candidate_limit": 40,
        "regions": ["Hong Kong", "Taiwan", "Mainland China", "Thailand", "Malaysia"],
        "target_titles": [
            "Chief Investment Officer",
            "CIO",
            "Family Office Director",
            "Head of Family Office",
            "Principal",
            "Managing Partner",
            "Investment Director",
            "Family Principal",
        ],
        "min_aum_usd": 200_000_000,
        "require_email": True,
        "min_email_confidence": 70,
    },
    "resources": {
        "apollo": True,
        "serper": True,
        "brave": True,
        "hunter": True,
        "gemini_scoring": True,
        "gemini_outreach": True,
    },
    "human_review": {
        "after_candidate_discovery": True,
        "after_research_enrichment": True,
        "before_outreach_generation": True,
    },
    "scoring": {
        "weights": {
            "match_score": 0.35,
            "reach_score": 0.25,
            "budget_score": 0.25,
            "cycle_score": 0.15,
        },
        "a_level_min": 30,
        "b_level_min": 20,
    },
    "rate_limit": {"request_delay_seconds": 1.2},
}

TASK_CONFIG = json.loads(json.dumps(DEFAULT_CONFIG))


@dataclass
class Lead:
    region: str = ""
    fo_name: str = ""
    domain: str = ""
    official_website: str = ""
    contact_name: str = ""
    contact_title: str = ""
    linkedin_url: str = ""
    email: str = ""
    phone: str = ""
    email_confidence: int = 0
    aum: str = "Undisclosed"
    aum_usd_estimate: Optional[int] = None
    core_business: str = ""
    core_pain: str = ""
    family_background: str = ""
    source_urls: str = ""
    match_score: int = 0
    reach_score: int = 0
    budget_score: int = 0
    cycle_score: int = 0
    total_score: int = 0
    customer_level: str = ""
    dev_suggest: str = ""
    outreach_email_en: str = ""
    personal_reachability: str = ""
    raw: Dict[str, Any] = field(default_factory=dict)

    def to_row(self) -> Dict[str, Any]:
        return {
            "region": self.region,
            "fo_name": self.fo_name,
            "domain": self.domain,
            "official_website": self.official_website,
            "contact_name": self.contact_name,
            "contact_title": self.contact_title,
            "linkedin_url": self.linkedin_url,
            "email": self.email,
            "phone": self.phone,
            "email_confidence": self.email_confidence,
            "aum": self.aum,
            "aum_usd_estimate": self.aum_usd_estimate or "",
            "core_business": self.core_business,
            "core_pain": self.core_pain,
            "family_background": self.family_background,
            "source_urls": self.source_urls,
            "match_score": self.match_score,
            "reach_score": self.reach_score,
            "budget_score": self.budget_score,
            "cycle_score": self.cycle_score,
            "total_score": self.total_score,
            "customer_level": self.customer_level,
            "dev_suggest": self.dev_suggest,
            "outreach_email_en": self.outreach_email_en,
            "personal_reachability": self.personal_reachability or describe_personal_reachability(self),
        }

    @classmethod
    def from_row(cls, row: Dict[str, Any]) -> "Lead":
        estimate = row.get("aum_usd_estimate") or None
        try:
            estimate = int(estimate) if estimate not in {"", None} else None
        except ValueError:
            estimate = None
        return cls(
            region=row.get("region", ""),
            fo_name=row.get("fo_name", ""),
            domain=row.get("domain", ""),
            official_website=row.get("official_website", "") or (
                f"https://{row.get('domain')}" if row.get("domain") else ""
            ),
            contact_name=row.get("contact_name", ""),
            contact_title=row.get("contact_title", ""),
            linkedin_url=row.get("linkedin_url", ""),
            email=row.get("email", ""),
            phone=row.get("phone", ""),
            email_confidence=int(row.get("email_confidence") or 0),
            aum=row.get("aum", "Undisclosed"),
            aum_usd_estimate=estimate,
            core_business=row.get("core_business", ""),
            core_pain=row.get("core_pain", ""),
            family_background=row.get("family_background", ""),
            source_urls=row.get("source_urls", ""),
            total_score=int(row.get("total_score") or 0),
            customer_level=row.get("customer_level", ""),
            personal_reachability=row.get("personal_reachability", ""),
        )


def env(name: str, required: bool = True) -> str:
    value = os.getenv(name, "").strip()
    if required and not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def load_config(path: str = "") -> Dict[str, Any]:
    config = json.loads(json.dumps(DEFAULT_CONFIG))
    if not path:
        return config
    user_config = json.loads(Path(path).read_text(encoding="utf-8"))
    deep_merge(config, user_config)
    return config


def deep_merge(base: Dict[str, Any], update: Dict[str, Any]) -> None:
    for key, value in update.items():
        if isinstance(value, dict) and isinstance(base.get(key), dict):
            deep_merge(base[key], value)
        else:
            base[key] = value


def cfg(path: str, default: Any = None) -> Any:
    cur: Any = TASK_CONFIG
    for part in path.split("."):
        if not isinstance(cur, dict) or part not in cur:
            return default
        cur = cur[part]
    return cur


def sleep() -> None:
    time.sleep(float(cfg("rate_limit.request_delay_seconds", 1.2)))


def post_json(url: str, headers: Dict[str, str], payload: Dict[str, Any]) -> Dict[str, Any]:
    resp = requests.post(url, headers=headers, json=payload, timeout=TIMEOUT)
    if resp.status_code >= 400:
        raise RuntimeError(f"POST {url} failed: {resp.status_code} {resp.text[:500]}")
    return resp.json()


def get_json(url: str, headers: Optional[Dict[str, str]] = None) -> Dict[str, Any]:
    resp = requests.get(url, headers=headers or {}, timeout=TIMEOUT)
    if resp.status_code >= 400:
        raise RuntimeError(f"GET {url} failed: {resp.status_code} {resp.text[:500]}")
    return resp.json()


def parse_region(location: str) -> str:
    loc = (location or "").lower()
    if "hong kong" in loc:
        return "香港"
    if "taiwan" in loc:
        return "台湾"
    if "china" in loc or "mainland" in loc or "shanghai" in loc or "beijing" in loc:
        return "大陆"
    if "thailand" in loc or "bangkok" in loc:
        return "泰国"
    if "malaysia" in loc or "kuala lumpur" in loc:
        return "马来西亚"
    return ""


def normalize_domain(domain: str) -> str:
    domain = (domain or "").strip().lower()
    domain = re.sub(r"^https?://", "", domain)
    domain = domain.strip("/")
    return domain


def company_domain(org: Dict[str, Any]) -> str:
    return normalize_domain(
        org.get("primary_domain")
        or org.get("website_url")
        or org.get("domain")
        or ""
    )


def fetch_sfo_candidates(limit: int) -> List[Lead]:
    if not cfg("resources.apollo", True):
        print("Apollo disabled by config; no candidates fetched.")
        return []
    api_key = env("APOLLO_API_KEY")
    url = "https://api.apollo.io/v1/mixed_people/search"
    headers = {
        "Content-Type": "application/json",
        "Cache-Control": "no-cache",
        "x-api-key": api_key,
    }
    payload = {
        "q_keywords": '"single family office" OR SFO OR "family office"',
        "person_titles": cfg("task.target_titles", DEFAULT_CONFIG["task"]["target_titles"]),
        "person_locations": cfg("task.regions", DEFAULT_CONFIG["task"]["regions"]),
        "page": 1,
        "per_page": min(max(limit, 1), 100),
    }
    data = post_json(url, headers, payload)
    people = data.get("people") or data.get("contacts") or []
    leads: List[Lead] = []

    for p in people:
        org = p.get("organization") or {}
        domain = company_domain(org)
        name = p.get("name") or " ".join(
            x for x in [p.get("first_name"), p.get("last_name")] if x
        )
        if not domain or not name:
            continue
        location = p.get("city") or p.get("state") or p.get("country") or p.get("location") or ""
        region = parse_region(str(location))
        lead = Lead(
            region=region,
            fo_name=org.get("name") or p.get("organization_name") or "",
            domain=domain,
            official_website=f"https://{domain}",
            contact_name=name,
            contact_title=p.get("title") or "",
            linkedin_url=p.get("linkedin_url") or "",
            phone=p.get("phone") or p.get("organization", {}).get("phone") or "",
            raw={"apollo": p},
        )
        leads.append(lead)
        if len(leads) >= limit:
            break
    return dedupe_leads(leads)


def dedupe_leads(leads: Iterable[Lead]) -> List[Lead]:
    seen = set()
    out: List[Lead] = []
    for lead in leads:
        key = (lead.domain.lower(), lead.contact_name.lower())
        if key in seen:
            continue
        seen.add(key)
        out.append(lead)
    return out


def combined_search_snippets(lead: Lead) -> Dict[str, str]:
    query = (
        f'"{lead.fo_name}" "{lead.contact_name}" single family office AUM '
        f'investment portfolio multi custodian reporting family background'
    )

    snippets: List[str] = []
    links: List[str] = []
    if cfg("resources.serper", True):
        serper_key = env("SERPER_API_KEY")
        serper_payload = {"q": query, "num": 8, "gl": "hk"}
        serper = post_json(
            "https://google.serper.dev/search",
            {"X-API-KEY": serper_key, "Content-Type": "application/json"},
            serper_payload,
        )
        sleep()
        for r in serper.get("organic", []):
            snippets.append(f"{r.get('title', '')}: {r.get('snippet', '')}")
            if r.get("link"):
                links.append(r["link"])
    if cfg("resources.brave", True):
        brave_key = env("BRAVE_API_KEY")
        brave_url = (
            "https://api.search.brave.com/res/v1/web/search"
            f"?q={quote_plus(query)}&count=8"
        )
        brave = get_json(brave_url, {"X-Subscription-Token": brave_key})
        sleep()
        for r in brave.get("web", {}).get("results", []):
            snippets.append(f"{r.get('title', '')}: {r.get('description', '')}")
            if r.get("url"):
                links.append(r["url"])

    return {"text": "\n".join(snippets)[:6000], "links": " | ".join(dict.fromkeys(links))}


def extract_json(text: str) -> Dict[str, Any]:
    cleaned = text.strip()
    cleaned = re.sub(r"^```(?:json)?", "", cleaned, flags=re.I).strip()
    cleaned = re.sub(r"```$", "", cleaned).strip()
    match = re.search(r"\{.*\}", cleaned, re.S)
    if not match:
        raise ValueError(f"No JSON object found in model output: {text[:300]}")
    return json.loads(match.group(0))


def gemini_generate(prompt: str, *, response_json: bool = False) -> str:
    api_key = env("GEMINI_API_KEY")
    model = os.getenv("GEMINI_MODEL", "gemini-1.5-flash-latest")
    url = (
        "https://generativelanguage.googleapis.com/v1beta/models/"
        f"{model}:generateContent?key={api_key}"
    )
    payload: Dict[str, Any] = {"contents": [{"parts": [{"text": prompt}]}]}
    if response_json:
        payload["generationConfig"] = {"response_mime_type": "application/json"}
    data = post_json(url, {"Content-Type": "application/json"}, payload)
    sleep()
    return data["candidates"][0]["content"]["parts"][0]["text"]


def enrich_fo_info(lead: Lead) -> None:
    search = combined_search_snippets(lead)
    prompt = f"""
Extract structured research for this single family office or single-family investment platform.
Return strict JSON only.

Organization: {lead.fo_name}
Domain: {lead.domain}
Contact: {lead.contact_name}, {lead.contact_title}
Search snippets:
{search["text"]}

Fields:
- aum: disclosed AUM in USD terms, or "Undisclosed"
- aum_usd_estimate: integer USD estimate if confidently inferable, otherwise null
- core_business: concise description
- core_pain: likely asset reporting/data operations pain relevant to Canopy
- family_background: family founder/background and cross-border asset footprint
- is_single_family_office: true/false/unknown
- evidence_note: one sentence on evidence quality
"""
    info = extract_json(gemini_generate(prompt, response_json=True))
    lead.aum = str(info.get("aum") or "Undisclosed")
    lead.aum_usd_estimate = info.get("aum_usd_estimate")
    lead.core_business = str(info.get("core_business") or "")
    lead.core_pain = str(info.get("core_pain") or "")
    lead.family_background = str(info.get("family_background") or "")
    lead.source_urls = search["links"]
    lead.raw["enrichment"] = info


def get_hunter_email(lead: Lead) -> None:
    if not cfg("resources.hunter", True):
        return
    if not lead.domain or not lead.contact_name:
        return
    api_key = env("HUNTER_API_KEY")
    url = (
        "https://api.hunter.io/v2/email-finder"
        f"?domain={quote_plus(lead.domain)}"
        f"&full_name={quote_plus(lead.contact_name)}"
        f"&api_key={quote_plus(api_key)}"
    )
    data = get_json(url)
    sleep()
    body = data.get("data") or {}
    email = body.get("email") or ""
    score = int(body.get("score") or body.get("confidence") or 0)
    status = ((body.get("verification") or {}).get("status") or "").lower()
    if email and score >= int(cfg("task.min_email_confidence", 70)) and status not in {"invalid"}:
        lead.email = email
        lead.email_confidence = score
    lead.raw["hunter"] = {"score": score, "status": status}


def score_fo(lead: Lead) -> None:
    if not cfg("resources.gemini_scoring", True):
        heuristic_score(lead)
        return
    prompt = f"""
任务：基于提供的单一家办全部信息，按照4个固定维度1-10分量化打分，仅输出标准JSON，无多余文字。

Canopy定位：面向单一家办和家族资本团队的资产聚合、投资组合报告、多托管行数据整合SaaS。

输入家办数据：
{json.dumps(lead.to_row(), ensure_ascii=False)}

定性判断框架：
- 适合Canopy的客户通常拥有跨银行/券商/外部管理人/私募基金/直接股权/地产等分散资产。
- 优先选择跨境、多币种、多实体、代际治理或投后报告压力明显的家族。
- 不适合优先开发：单一产业控股为主、缺少外部金融资产证据、联系人弱、公开资料无法验证的对象。

四大评估维度打分规则（1分最低，10分最高）：
1. match_score 产品匹配度：考核AUM规模、跨境资产数量、多托管行需求与Canopy适配程度。
2. reach_score 触达可行性：考核联系方式完整度、联系人是否拥有采购决策权。
   重要：公司官网表单、公司总机、info@邮箱、公司LinkedIn主页不算个人可触达；只有个人商务邮箱、个人LinkedIn、个人直线/办公分机才计入高分。
3. budget_score 付费潜力：考核家族财富科技预算、替换现有系统意愿。
4. cycle_score 转化周期：考核家族决策链条长短、数字化改造迫切度。

附加输出字段：
1. total_score：四项分数相加总和。
2. customer_level：A类高优客户=总分≥30；B类培育客户=20≤总分＜30；C类低优先级=总分＜20。
3. dev_suggest：简短开发动作建议，20字以内。
"""
    score = extract_json(gemini_generate(prompt, response_json=True))
    lead.match_score = int(score.get("match_score") or 0)
    lead.reach_score = int(score.get("reach_score") or 0)
    lead.budget_score = int(score.get("budget_score") or 0)
    lead.cycle_score = int(score.get("cycle_score") or 0)
    lead.total_score = int(score.get("total_score") or (
        lead.match_score + lead.reach_score + lead.budget_score + lead.cycle_score
    ))
    lead.customer_level = str(score.get("customer_level") or "")
    lead.dev_suggest = str(score.get("dev_suggest") or "")


def heuristic_score(lead: Lead) -> None:
    match = 4
    text = " ".join([lead.core_business, lead.core_pain, lead.family_background]).lower()
    for token in ["cross-border", "global", "multi-currency", "multi custodian", "private equity", "fund"]:
        if token in text:
            match += 1
    reach = 3 + bool(lead.linkedin_url) * 2 + bool(lead.email) * 3 + bool(lead.phone) * 1
    budget = 6 + bool(lead.aum_usd_estimate and lead.aum_usd_estimate >= 500_000_000) * 2
    cycle = 5 + bool(lead.email) * 1 + bool("manual" in text or "report" in text) * 1
    lead.match_score = min(match, 10)
    lead.reach_score = min(reach, 10)
    lead.budget_score = min(budget, 10)
    lead.cycle_score = min(cycle, 10)
    lead.total_score = lead.match_score + lead.reach_score + lead.budget_score + lead.cycle_score
    if lead.total_score >= int(cfg("scoring.a_level_min", 30)):
        lead.customer_level = "A类高优客户"
    elif lead.total_score >= int(cfg("scoring.b_level_min", 20)):
        lead.customer_level = "B类培育客户"
    else:
        lead.customer_level = "C类低优先级"
    lead.dev_suggest = "人工复核后触达"


def generate_outreach_email(lead: Lead) -> None:
    if not cfg("resources.gemini_outreach", True):
        return
    prompt = f"""
身份设定：Canopy亚太区域财富科技BD，面向亚洲单一家办决策人撰写英文开发邮件。

输入变量：
{json.dumps(lead.to_row(), ensure_ascii=False)}

知识库参考：
Canopy聚合325+全球托管行数据，支持自动多币种合并报表、实时资产视图、另类资产和直接投资记录，帮助家族办公室减少手工对账和Excel报表工作。

硬性要求：
1. 全文英文商务正式文风，总词数严格控制在190-210词。
2. 开头必须结合对方家族背景/区域业务做个性化切入，禁止通用模板。
3. 精准匹配其资产痛点，对比传统Excel、多银行独立后台的低效问题。
4. 插入1个同区域、同等AUM的匿名SFO落地成功案例，不要编造客户名称。
5. 清晰阐述Canopy核心价值：聚合325+全球托管行、自动多币种合并报表、实时资产视图、减少80%手工对账工时。
6. 温和邀约15分钟免费线上产品演示，留下沟通预期。
7. 结尾标准BD商务署名，只输出邮件正文，不要任何解释、标题、备注。
"""
    lead.outreach_email_en = gemini_generate(prompt).strip()


def should_keep(lead: Lead, require_email: bool) -> bool:
    if require_email and not lead.email:
        return False
    if not has_personal_reachability(lead):
        return False
    min_aum = int(cfg("task.min_aum_usd", 200_000_000))
    if lead.aum_usd_estimate is not None and lead.aum_usd_estimate < min_aum:
        return False
    return True


def has_personal_reachability(lead: Lead) -> bool:
    if is_personal_business_email(lead.email):
        return True
    if lead.linkedin_url and "/in/" in lead.linkedin_url.lower():
        return True
    if lead.phone and lead.contact_name:
        return True
    return False


def describe_personal_reachability(lead: Lead) -> str:
    methods: List[str] = []
    if is_personal_business_email(lead.email):
        methods.append("personal_business_email")
    if lead.linkedin_url and "/in/" in lead.linkedin_url.lower():
        methods.append("personal_linkedin")
    if lead.phone and lead.contact_name:
        methods.append("personal_phone_or_direct_office_line")
    return ";".join(methods) if methods else "not_personally_reachable"


def is_personal_business_email(email: str) -> bool:
    if not email or "@" not in email:
        return False
    local = email.split("@", 1)[0].lower()
    generic = {
        "info",
        "contact",
        "enquiry",
        "enquiries",
        "hello",
        "admin",
        "office",
        "mail",
        "support",
        "service",
        "services",
        "marketing",
        "media",
        "pr",
        "press",
        "investor",
        "investors",
        "ir",
        "careers",
        "hr",
        "noreply",
        "no-reply",
    }
    return local not in generic


def read_review_file(path: str) -> List[Lead]:
    leads: List[Lead] = []
    with Path(path).open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            approve = (row.get("approve") or "").strip().upper()
            if approve not in {"Y", "YES", "TRUE", "1"}:
                continue
            leads.append(Lead.from_row(row))
    return leads


def write_outputs(leads: List[Lead], out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    csv_path = out_dir / f"canopy_sfo_validated_leads_{today}.csv"
    json_path = out_dir / f"canopy_sfo_validated_leads_{today}.json"
    md_path = out_dir / f"canopy_sfo_evaluation_workflow_{today}.md"

    rows = [lead.to_row() for lead in leads]
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()) if rows else [])
        writer.writeheader()
        writer.writerows(rows)
    json_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    md_path.write_text(build_workflow_doc(leads), encoding="utf-8")
    maybe_write_excel(rows, out_dir / f"canopy_sfo_validated_leads_{today}.xlsx")
    maybe_write_manifest(out_dir, [csv_path, json_path, md_path])


def write_review_file(leads: List[Lead], out_dir: Path, stage: str) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"review_{stage}.csv"
    fields = [
        "approve",
        "review_note",
        "region",
        "fo_name",
        "domain",
        "contact_name",
        "contact_title",
        "linkedin_url",
        "email",
        "phone",
        "aum",
        "aum_usd_estimate",
        "core_business",
        "core_pain",
        "family_background",
        "source_urls",
        "personal_reachability",
        "total_score",
        "customer_level",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for lead in leads:
            row = lead.to_row()
            row["approve"] = "Y"
            row["review_note"] = ""
            writer.writerow({field: row.get(field, "") for field in fields})
    return path


def stop_for_review(leads: List[Lead], out_dir: Path, stage: str) -> None:
    path = write_review_file(leads, out_dir, stage)
    print(f"Human review checkpoint written: {path}")
    print("Review the file, set approve to Y/N, then rerun the next stage if needed.")


def review_enabled(stage: str) -> bool:
    return bool(cfg(f"human_review.{stage}", False))


def maybe_write_excel(rows: List[Dict[str, Any]], xlsx_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "SFO Leads"
    if not rows:
        wb.save(xlsx_path)
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for idx, header in enumerate(headers, start=1):
        width = min(max(len(header) + 2, 14), 42)
        ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(xlsx_path)


def maybe_write_manifest(out_dir: Path, paths: List[Path]) -> None:
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "files": {
            p.name: hashlib.sha256(p.read_bytes()).hexdigest()
            for p in paths
            if p.exists()
        },
    }
    secret = os.getenv("JWT_SECRET", "").strip()
    if secret:
        try:
            import jwt
            payload["jwt"] = jwt.encode(payload, secret, algorithm="HS256")
        except Exception as exc:
            payload["jwt_error"] = f"PyJWT unavailable or signing failed: {exc}"
    (out_dir / "run_manifest.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def build_workflow_doc(leads: List[Lead]) -> str:
    regions = ", ".join(cfg("task.regions", DEFAULT_CONFIG["task"]["regions"]))
    min_aum = int(cfg("task.min_aum_usd", 200_000_000))
    return f"""# Canopy SFO Prospecting Workflow

Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}

## Scope

Regions: {regions}.
Target: single family offices or single-family investment platforms with AUM >= USD {min_aum:,} where public evidence supports a strong Canopy fit.

## API Roles

- Apollo: candidate discovery for SFO keywords, target titles, company domains, LinkedIn profiles.
- Serper + Brave Search: dual-engine public research for official sites, media, AUM, family background, cross-border assets and reporting pain points.
- Hunter.io: business email discovery from verified company domain plus contact name.
- Gemini: structured extraction, four-dimension scoring and personalized 190-210 word English outreach email.
- JWT secret: optional run manifest signing for internal audit integrity.

## Filtering Logic

1. Pull up to the requested candidate count from Apollo.
2. Human review checkpoint: remove MFOs, service providers, weak contacts and irrelevant group-company records.
3. Enrich every approved candidate with dual-search evidence.
4. Human review checkpoint: verify source quality, family-office fit, AUM evidence and likely Canopy pain.
5. Use Hunter Email Finder and keep only high-confidence business emails when configured.
6. Score remaining prospects and sort by total score.
7. Human review checkpoint: approve final account/contact list before outreach generation.
8. Generate outreach emails only for approved leads and export the top target count.

## Scoring

Each dimension is 1-10:

- match_score: cross-border, multi-custodian, multi-asset fit for Canopy.
- reach_score: email, LinkedIn, phone and decision-maker quality.
- budget_score: likely wealth-tech budget and ability to buy.
- cycle_score: urgency and likely speed of internal decision.

Qualitative account-fit signals:

- Strong fit: multi-bank custody, external managers, private funds, direct deals, real estate, cross-border structures, family governance and recurring reporting pressure.
- Weak fit: mostly operating-company revenue, no evidence of investable financial assets, no clear family capital team, generic corporate contact only.
- Research priority: where public data is incomplete but the family has obvious asset complexity, use warm-introduction channels before direct outreach.

Customer level:

- A: total score >= 30.
- B: total score 20-29.
- C: total score < 20.

## Results Summary

- Exported leads: {len(leads)}
- A-level: {sum(1 for x in leads if x.customer_level.startswith("A"))}
- B-level: {sum(1 for x in leads if x.customer_level.startswith("B"))}
- C-level: {sum(1 for x in leads if x.customer_level.startswith("C"))}

## Compliance Notes

The script does not guess private emails or phone numbers. It only keeps Hunter-returned business emails above the configured confidence threshold. All model-generated facts should be checked against the `source_urls` column before sending outreach.
"""


def run(args: argparse.Namespace) -> None:
    global TASK_CONFIG
    TASK_CONFIG = load_config(args.config)
    if args.target is not None:
        TASK_CONFIG["task"]["target_count"] = args.target
    if args.candidate_limit is not None:
        TASK_CONFIG["task"]["candidate_limit"] = args.candidate_limit
    if args.require_email:
        TASK_CONFIG["task"]["require_email"] = True

    out_dir = Path(args.out_dir)
    target = int(cfg("task.target_count", 30))
    candidate_limit = int(cfg("task.candidate_limit", 40))
    require_email = bool(cfg("task.require_email", True))

    stage = args.stage

    if stage == "discover":
        leads = fetch_sfo_candidates(candidate_limit)
        print(f"Fetched candidates: {len(leads)}")
        stop_for_review(leads, out_dir, "candidate_discovery")
        return

    if args.review_input:
        leads = read_review_file(args.review_input)
        print(f"Loaded approved review leads: {len(leads)}")
    else:
        leads = fetch_sfo_candidates(candidate_limit)
        print(f"Fetched candidates: {len(leads)}")
        if stage in {"full", "research"} and review_enabled("after_candidate_discovery"):
            stop_for_review(leads, out_dir, "candidate_discovery")
            return

    if stage == "outreach":
        outreach_ready: List[Lead] = []
        for lead in leads:
            if not has_personal_reachability(lead):
                print(f"  skipped not personally reachable: {lead.fo_name} - {lead.contact_name}")
                continue
            generate_outreach_email(lead)
            outreach_ready.append(lead)
        write_outputs(outreach_ready, out_dir)
        print(f"Exported leads: {len(outreach_ready)} -> {out_dir}")
        return

    validated: List[Lead] = []
    for idx, lead in enumerate(leads, start=1):
        print(f"[{idx}/{len(leads)}] {lead.fo_name} - {lead.contact_name}")
        try:
            enrich_fo_info(lead)
            get_hunter_email(lead)
            if not should_keep(lead, require_email):
                continue
            score_fo(lead)
            validated.append(lead)
            validated.sort(key=lambda x: x.total_score, reverse=True)
            validated = validated[: target]
        except Exception as exc:
            print(f"  skipped: {exc}", file=sys.stderr)
        if len(validated) >= target:
            break

    if stage in {"full", "research"} and review_enabled("after_research_enrichment"):
        stop_for_review(validated, out_dir, "research_enrichment")
        return

    if stage in {"full", "research"} and review_enabled("before_outreach_generation"):
        stop_for_review(validated, out_dir, "before_outreach_generation")
        return

    for lead in validated:
        try:
            generate_outreach_email(lead)
        except Exception as exc:
            print(f"  outreach skipped for {lead.fo_name}: {exc}", file=sys.stderr)

    validated.sort(key=lambda x: x.total_score, reverse=True)
    write_outputs(validated, out_dir)
    print(f"Exported leads: {len(validated)} -> {out_dir}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--stage",
        choices=["discover", "research", "outreach", "full"],
        default="full",
        help="Run a specific semi-automated stage.",
    )
    parser.add_argument("--config", default="", help="Path to task config JSON.")
    parser.add_argument("--target", type=int, default=None)
    parser.add_argument("--candidate-limit", type=int, default=None)
    parser.add_argument("--out-dir", default="output")
    parser.add_argument("--review-input", default="", help="Approved review CSV from a prior checkpoint.")
    parser.add_argument(
        "--require-email",
        action="store_true",
        help="Keep only leads with Hunter high-confidence business email.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    run(parse_args())
